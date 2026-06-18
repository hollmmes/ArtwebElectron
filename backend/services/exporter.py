import csv
import json
import io
import xml.etree.ElementTree as ET
from datetime import datetime


FIELDS = [
    'name', 'category', 'address', 'phone', 'website', 'emails',
    'rating', 'reviews_count', 'reviews', 'working_hours', 'about',
    'facebook', 'instagram', 'twitter', 'youtube', 'linkedin', 'tiktok',
    'maps_url', 'latitude', 'longitude'
]

FIELD_LABELS = {
    'name': 'İşletme Adı',
    'category': 'Kategori',
    'address': 'Adres',
    'phone': 'Telefon',
    'website': 'Website',
    'emails': 'E-posta',
    'rating': 'Puan',
    'reviews_count': 'Yorum Sayısı',
    'reviews': 'Yorumlar',
    'working_hours': 'Çalışma Saatleri',
    'about': 'Hakkında',
    'facebook': 'Facebook',
    'instagram': 'Instagram',
    'twitter': 'Twitter/X',
    'youtube': 'YouTube',
    'linkedin': 'LinkedIn',
    'tiktok': 'TikTok',
    'maps_url': 'Google Maps',
    'latitude': 'Enlem',
    'longitude': 'Boylam',
}


def _build_row(biz: dict) -> dict:
    social = biz.get('social_media', {}) or {}
    emails = biz.get('emails', [])
    return {
        'name': biz.get('name', '') or '',
        'category': biz.get('category', '') or '',
        'address': biz.get('address', '') or '',
        'phone': biz.get('phone', '') or '',
        'website': biz.get('website', '') or '',
        'emails': '; '.join(emails) if isinstance(emails, list) else (emails or ''),
        'rating': biz.get('rating', '') if biz.get('rating') is not None else '',
        'reviews_count': biz.get('reviews_count', '') if biz.get('reviews_count') is not None else '',
        'reviews': _format_reviews(biz.get('reviews', [])),
        'working_hours': _format_hours(biz.get('working_hours', {})),
        'about': biz.get('about', '') or '',
        'facebook': social.get('facebook', '') or '',
        'instagram': social.get('instagram', '') or '',
        'twitter': social.get('twitter', '') or '',
        'youtube': social.get('youtube', '') or '',
        'linkedin': social.get('linkedin', '') or '',
        'tiktok': social.get('tiktok', '') or '',
        'maps_url': biz.get('maps_url', '') or '',
        'latitude': biz.get('latitude', '') if biz.get('latitude') is not None else '',
        'longitude': biz.get('longitude', '') if biz.get('longitude') is not None else '',
    }


def export_to_csv(businesses: list[dict]) -> bytes:
    if not businesses:
        return ''.encode('utf-8-sig')

    output = io.StringIO()
    writer = csv.DictWriter(output, fieldnames=FIELDS, extrasaction='ignore')
    writer.writerow({f: FIELD_LABELS.get(f, f) for f in FIELDS})
    for biz in businesses:
        writer.writerow(_build_row(biz))

    return output.getvalue().encode('utf-8-sig')


def export_to_json(businesses: list[dict]) -> bytes:
    clean = []
    for biz in businesses:
        clean.append({
            'name': biz.get('name', ''),
            'category': biz.get('category', ''),
            'address': biz.get('address', ''),
            'phone': biz.get('phone', ''),
            'website': biz.get('website', ''),
            'emails': biz.get('emails', []),
            'rating': biz.get('rating'),
            'reviews_count': biz.get('reviews_count'),
            'working_hours': biz.get('working_hours', {}),
            'about': biz.get('about', ''),
            'social_media': biz.get('social_media', {}),
            'maps_url': biz.get('maps_url', ''),
            'latitude': biz.get('latitude'),
            'longitude': biz.get('longitude'),
            'reviews': biz.get('reviews', []),
        })
    return json.dumps(clean, ensure_ascii=False, indent=2).encode('utf-8')


def export_to_xml(businesses: list[dict]) -> bytes:
    root = ET.Element('isletmeler')
    root.set('olusturulma', datetime.now().strftime('%Y-%m-%d %H:%M'))
    root.set('toplam', str(len(businesses)))

    for biz in businesses:
        row = _build_row(biz)
        item = ET.SubElement(root, 'isletme')
        for field in FIELDS:
            child = ET.SubElement(item, field)
            child.text = str(row.get(field, ''))

    tree = ET.ElementTree(root)
    buf = io.BytesIO()
    tree.write(buf, encoding='utf-8', xml_declaration=True)
    return buf.getvalue()


def export_to_xlsx(businesses: list[dict]) -> bytes:
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, GradientFill
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise RuntimeError("openpyxl kurulu değil. 'pip install openpyxl' ile kurun.")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'İşletmeler'

    # Renk paleti
    HEADER_BG   = '1E3A5F'   # koyu lacivert
    HEADER_FG   = 'FFFFFF'
    ALT_ROW_BG  = 'EEF4FB'   # açık mavi-gri
    LINK_COLOR  = '2563EB'
    BORDER_COLOR = 'D1D5DB'

    thin = Side(style='thin', color=BORDER_COLOR)
    cell_border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # ── Başlık satırı ──────────────────────────────────────────────
    header_font  = Font(bold=True, color=HEADER_FG, size=10, name='Calibri')
    header_fill  = PatternFill(start_color=HEADER_BG, end_color=HEADER_BG, fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)

    headers = [FIELD_LABELS.get(f, f) for f in FIELDS]
    ws.append(headers)
    ws.row_dimensions[1].height = 36

    for col_idx, cell in enumerate(ws[1], 1):
        cell.font   = cell.font = header_font
        cell.fill   = header_fill
        cell.alignment = header_align
        cell.border = cell_border

    # ── Veri satırları ─────────────────────────────────────────────
    url_fields = {'website', 'maps_url', 'facebook', 'instagram', 'twitter', 'youtube', 'linkedin', 'tiktok'}
    alt_fill = PatternFill(start_color=ALT_ROW_BG, end_color=ALT_ROW_BG, fill_type='solid')
    link_font  = Font(color=LINK_COLOR, underline='single', size=10, name='Calibri')
    normal_font = Font(size=10, name='Calibri')
    wrap_align  = Alignment(vertical='top', wrap_text=True)
    nowrap_align = Alignment(vertical='top', wrap_text=False)

    for row_idx, biz in enumerate(businesses, 2):
        row = _build_row(biz)
        is_alt = (row_idx % 2 == 0)
        row_fill = alt_fill if is_alt else None
        ws.row_dimensions[row_idx].height = 52  # URL'ler için biraz yüksek

        for col_idx, field in enumerate(FIELDS, 1):
            val = row.get(field, '')
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = cell_border

            if row_fill:
                cell.fill = row_fill

            if field in url_fields and val and val.startswith('http'):
                cell.hyperlink = val
                cell.value = val
                cell.font = link_font
                cell.alignment = nowrap_align
            elif field in ('working_hours', 'about', 'address', 'reviews'):
                cell.font = normal_font
                cell.alignment = wrap_align
            else:
                cell.font = normal_font
                cell.alignment = nowrap_align

    # ── Kolon genişlikleri ─────────────────────────────────────────
    col_widths = {
        'name': 32, 'category': 20, 'address': 38, 'phone': 16,
        'website': 28, 'emails': 28, 'rating': 8, 'reviews_count': 10,
        'reviews': 60, 'working_hours': 34, 'about': 38,
        'facebook': 28, 'instagram': 28, 'twitter': 24,
        'youtube': 28, 'linkedin': 28, 'tiktok': 24,
        'maps_url': 28, 'latitude': 12, 'longitude': 12,
    }
    for col_idx, field in enumerate(FIELDS, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = col_widths.get(field, 18)

    # ── Freeze & filtre ────────────────────────────────────────────
    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = ws.dimensions

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def export_to_html(businesses: list[dict], query: str = '', location: str = '') -> bytes:
    now = datetime.now().strftime('%d.%m.%Y %H:%M')
    title = f"İşletme Raporu — {query}" if query else "İşletme Raporu"

    # Görünmesini istemediğimiz teknik alanlar
    VISIBLE_FIELDS = [
        'name', 'category', 'address', 'phone', 'website', 'emails',
        'rating', 'reviews_count', 'working_hours', 'about',
        'facebook', 'instagram', 'twitter', 'youtube', 'linkedin', 'tiktok',
        'maps_url',
    ]
    url_fields = {'website', 'maps_url', 'facebook', 'instagram', 'twitter', 'youtube', 'linkedin', 'tiktok'}

    def cell_html(field, val):
        if not val:
            return '<span class="empty">—</span>'
        if field in url_fields and str(val).startswith('http'):
            short = str(val).replace('https://', '').replace('http://', '').rstrip('/')
            if len(short) > 35:
                short = short[:33] + '…'
            return f'<a href="{val}" target="_blank">{short}</a>'
        if field == 'rating' and val:
            return f'<span class="rating">★ {val}</span>'
        if field == 'reviews_count' and val:
            return f'<span class="reviews">{val} yorum</span>'
        # Satır sonlarını koru
        return str(val).replace('\n', '<br>')

    rows_html = ''
    for i, biz in enumerate(businesses):
        row = _build_row(biz)
        cells = ''.join(
            f'<td class="field-{f}">{cell_html(f, row.get(f,""))}</td>'
            for f in VISIBLE_FIELDS
        )
        rows_html += f'<tr class="{"alt" if i % 2 else ""}">{cells}</tr>\n'

    header_cells = ''.join(
        f'<th>{FIELD_LABELS.get(f, f)}</th>' for f in VISIBLE_FIELDS
    )

    subtitle = f'{query}' + (f' · {location}' if location else '') + f' · {len(businesses)} işletme · {now}'

    html = f'''<!DOCTYPE html>
<html lang="tr">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>{title}</title>
<style>
  *, *::before, *::after {{ box-sizing: border-box; margin: 0; padding: 0; }}
  body {{ font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", Roboto, sans-serif;
          background: #f8fafc; color: #1e293b; font-size: 13px; }}

  .page-header {{
    background: linear-gradient(135deg, #1e3a5f 0%, #2563eb 100%);
    color: #fff; padding: 28px 40px 24px;
  }}
  .page-header h1 {{ font-size: 22px; font-weight: 700; margin-bottom: 4px; }}
  .page-header p  {{ font-size: 12px; opacity: .75; }}

  .stats {{
    display: flex; gap: 16px; padding: 16px 40px;
    background: #fff; border-bottom: 1px solid #e2e8f0;
  }}
  .stat {{ background: #f1f5f9; border-radius: 8px; padding: 10px 20px; text-align: center; }}
  .stat-val {{ font-size: 20px; font-weight: 700; color: #2563eb; }}
  .stat-lbl {{ font-size: 11px; color: #64748b; margin-top: 2px; }}

  .table-wrap {{ padding: 24px 40px 48px; overflow-x: auto; }}

  table {{ width: 100%; border-collapse: collapse; background: #fff;
           border-radius: 10px; overflow: hidden;
           box-shadow: 0 1px 3px rgba(0,0,0,.08), 0 0 0 1px rgba(0,0,0,.04); }}

  thead tr {{ background: #1e3a5f; color: #fff; }}
  th {{ padding: 11px 14px; text-align: left; font-size: 11px; font-weight: 600;
        letter-spacing: .04em; text-transform: uppercase; white-space: nowrap; }}

  td {{ padding: 10px 14px; border-bottom: 1px solid #f1f5f9; vertical-align: top;
        line-height: 1.5; }}
  tr.alt td {{ background: #f8fbff; }}
  tr:last-child td {{ border-bottom: none; }}
  tr:hover td {{ background: #eff6ff !important; }}

  /* Alan genişlikleri */
  .field-name      {{ min-width: 160px; font-weight: 600; color: #0f172a; }}
  .field-category  {{ min-width: 110px; }}
  .field-address   {{ min-width: 200px; max-width: 260px; }}
  .field-phone     {{ min-width: 110px; white-space: nowrap; }}
  .field-website, .field-maps_url,
  .field-facebook, .field-instagram, .field-twitter,
  .field-youtube,  .field-linkedin,  .field-tiktok
                   {{ min-width: 140px; max-width: 180px; overflow: hidden; }}
  .field-emails    {{ min-width: 160px; }}
  .field-rating    {{ min-width: 64px; white-space: nowrap; }}
  .field-reviews_count {{ min-width: 80px; white-space: nowrap; }}
  .field-working_hours {{ min-width: 180px; max-width: 240px; }}
  .field-about     {{ min-width: 200px; max-width: 300px; }}

  a {{ color: #2563eb; text-decoration: none; word-break: break-all; }}
  a:hover {{ text-decoration: underline; }}

  .rating  {{ color: #d97706; font-weight: 700; }}
  .reviews {{ color: #64748b; font-size: 12px; }}
  .empty   {{ color: #cbd5e1; }}

  .footer {{ text-align: center; padding: 16px; color: #94a3b8; font-size: 11px;
             border-top: 1px solid #e2e8f0; background: #fff; }}

  @media print {{
    body {{ background: #fff; }}
    .page-header {{ background: #1e3a5f !important; -webkit-print-color-adjust: exact; print-color-adjust: exact; }}
    thead tr {{ background: #1e3a5f !important; -webkit-print-color-adjust: exact; print-color-adjust: exact; }}
    .table-wrap {{ padding: 12px; }}
    .stats {{ padding: 12px; }}
  }}
</style>
</head>
<body>

<div class="page-header">
  <h1>{title}</h1>
  <p>{subtitle}</p>
</div>

<div class="stats">
  <div class="stat"><div class="stat-val">{len(businesses)}</div><div class="stat-lbl">İşletme</div></div>
  <div class="stat"><div class="stat-val">{sum(1 for b in businesses if b.get("phone"))}</div><div class="stat-lbl">Telefon</div></div>
  <div class="stat"><div class="stat-val">{sum(1 for b in businesses if b.get("website"))}</div><div class="stat-lbl">Website</div></div>
  <div class="stat"><div class="stat-val">{sum(1 for b in businesses if b.get("emails"))}</div><div class="stat-lbl">E-posta</div></div>
</div>

<div class="table-wrap">
  <table>
    <thead><tr>{header_cells}</tr></thead>
    <tbody>{rows_html}</tbody>
  </table>
</div>

<div class="footer">Art Web Toolkit &nbsp;·&nbsp; {now}</div>

</body>
</html>'''

    return html.encode('utf-8')


def _format_hours(hours: dict) -> str:
    if not hours:
        return ''
    if 'info' in hours:
        return hours['info']
    return '; '.join(f"{day}: {time}" for day, time in hours.items())


def _format_reviews(reviews: list) -> str:
    if not reviews:
        return ''
    parts = []
    for r in reviews:
        author = r.get('author', '') or ''
        rating = r.get('rating')
        text = r.get('text', '') or ''
        line = f"[{author}"
        if rating:
            line += f" ★{rating}"
        line += "]"
        if text:
            line += f" {text}"
        parts.append(line)
    return '\n'.join(parts)
